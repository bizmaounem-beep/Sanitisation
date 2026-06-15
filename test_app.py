import os
import sys
import base64
import tempfile
import unittest
import json
import sqlite3
import datetime
from app import app, DATABASE, ACTIVE_SESSIONS, generate_password_hash
from app import limiter

class SanitationSystemTestCase(unittest.TestCase):
    def setUp(self):
        # Configure app for testing
        app.config['TESTING'] = True
        app.config['RATELIMIT_ENABLED'] = False  # Disable rate limiting during tests
        limiter.enabled = False
        self.client = app.test_client()
        
        # Override database to use a temporary file
        self.db_fd, self.db_path = tempfile.mkstemp()
        import app as app_module
        app_module.DATABASE = self.db_path
        
        # Initialize database schema and seeds
        with app.app_context():
            db = sqlite3.connect(self.db_path)
            db.row_factory = sqlite3.Row
            with open('schema.sql', 'r', encoding='utf-8') as f:
                db.executescript(f.read())
            db.commit()
            
            # Auto hash seeds
            cursor = db.cursor()
            cursor.execute("SELECT id, password FROM users")
            users = cursor.fetchall()
            for u in users:
                user_id = u['id']
                raw_pwd = u['password']
                if not (raw_pwd.startswith('pbkdf2:') or raw_pwd.startswith('scrypt:')):
                    hashed = generate_password_hash(raw_pwd)
                    cursor.execute("UPDATE users SET password = ? WHERE id = ?", (hashed, user_id))
            db.commit()
            db.close()
            
    def tearDown(self):
        ACTIVE_SESSIONS.clear()
        import gc
        gc.collect()
        try:
            os.close(self.db_fd)
        except OSError:
            pass
        try:
            os.unlink(self.db_path)
        except PermissionError:
            pass
        
    def test_login_credentials(self):
        response = self.client.post('/api/auth/login', json={
            'username': 'admin',
            'password': 'admin'
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('token', data)
        self.assertEqual(data['user']['username'], 'admin')
        self.assertEqual(data['user']['role'], 'coordinator')
        
    def test_login_worker_credentials(self):
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'password'
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('token', data)
        self.assertEqual(data['user']['username'], 'worker1')
        self.assertEqual(data['user']['role'], 'worker')
        
    def test_register_supervisor_pending(self):
        response = self.client.post('/api/auth/register', json={
            'username': 'supervisor_new',
            'password': 'password123',
            'name': 'New Supervisor',
            'role': 'supervisor'
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertTrue(data['pending_approval'])
        
    def test_worker_password_recovery(self):
        # 1. Submit reset request
        response = self.client.post('/api/auth/request_reset', json={
            'username': 'worker1',
            'name': 'Waleed Worker'
        })
        self.assertEqual(response.status_code, 200)

        # Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        admin_token = json.loads(resp.data)['token']
        
        # 2. Verify reset_requested is 1 in user list
        users_resp = self.client.get('/api/users', headers={'Authorization': f'Bearer {admin_token}'})
        users = json.loads(users_resp.data)
        worker_user = next(u for u in users if u['username'] == 'worker1')
        self.assertEqual(worker_user['reset_requested'], 1)

        # 3. Coordinator generates reset code for worker1 (id=3)
        response = self.client.post('/api/users/3/reset_code', headers={'Authorization': f'Bearer {admin_token}'})
        self.assertEqual(response.status_code, 200)
        reset_code_worker = json.loads(response.data)['reset_code']

        # 4. Verify reset_requested is back to 0
        users_resp = self.client.get('/api/users', headers={'Authorization': f'Bearer {admin_token}'})
        users = json.loads(users_resp.data)
        worker_user = next(u for u in users if u['username'] == 'worker1')
        self.assertEqual(worker_user['reset_requested'], 0)

        # 5. Successful recovery verification for worker using the reset code
        response = self.client.post('/api/auth/recover', json={
            'username': 'worker1',
            'name': 'Waleed Worker',
            'reset_code': reset_code_worker
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(data['status'], 'success')
        self.assertEqual(data['role'], 'worker')
        self.assertTrue(data['requires_reset'])

        # 6. Reset worker password
        response = self.client.post('/api/auth/reset_password', json={
            'username': 'worker1',
            'name': 'Waleed Worker',
            'reset_code': reset_code_worker,
            'new_password': 'newworkerpass123'
        })
        self.assertEqual(response.status_code, 200)

        # 7. Verify standard login with the new password
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'newworkerpass123'
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertIn('token', data)
        
    def test_supervisor_password_reset(self):
        # Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        admin_token = json.loads(resp.data)['token']
        
        # Coordinator generates reset code for supervisor1 (id=2)
        response = self.client.post('/api/users/2/reset_code', headers={'Authorization': f'Bearer {admin_token}'})
        self.assertEqual(response.status_code, 200)
        reset_code_sup = json.loads(response.data)['reset_code']

        # 1. Verify identity requires password reset
        response = self.client.post('/api/auth/recover', json={
            'username': 'supervisor1',
            'name': 'Sarah Supervisor',
            'reset_code': reset_code_sup
        })
        data = json.loads(response.data)
        self.assertEqual(response.status_code, 200)
        self.assertEqual(data['status'], 'success')
        self.assertTrue(data['requires_reset'])
        
        # 2. Reset the password using the reset code
        response = self.client.post('/api/auth/reset_password', json={
            'username': 'supervisor1',
            'name': 'Sarah Supervisor',
            'reset_code': reset_code_sup,
            'new_password': 'newpassword123'
        })
        self.assertEqual(response.status_code, 200)
        
        # 3. Test standard login with the new password
        login_response = self.client.post('/api/auth/login', json={
            'username': 'supervisor1',
            'password': 'newpassword123'
        })
        self.assertEqual(login_response.status_code, 200)
        
    def test_recovery_invalid_user(self):
        # 1. Invalid username/name combination
        response = self.client.post('/api/auth/recover', json={
            'username': 'worker1',
            'name': 'Wrong Name',
            'reset_code': '123456'
        })
        self.assertEqual(response.status_code, 404)
        
        # 2. Missing parameter
        response = self.client.post('/api/auth/recover', json={
            'username': 'worker1'
        })
        self.assertEqual(response.status_code, 400)

    def test_generate_reset_code_unauthorized(self):
        # Login as supervisor
        resp = self.client.post('/api/auth/login', json={'username': 'supervisor1', 'password': 'password'})
        sup_token = json.loads(resp.data)['token']
        
        # Supervisor tries to generate reset code for worker1 (id=3)
        response = self.client.post('/api/users/3/reset_code', headers={'Authorization': f'Bearer {sup_token}'})
        self.assertEqual(response.status_code, 403)

    def test_task_workflow_and_consumption(self):
        # 1. Login as supervisor
        resp = self.client.post('/api/auth/login', json={'username': 'supervisor1', 'password': 'password'})
        sup_token = json.loads(resp.data)['token']
        
        # 2. Login as worker using credentials
        resp = self.client.post('/api/auth/login', json={'username': 'worker1', 'password': 'password'})
        work_token = json.loads(resp.data)['token']
        
        # 3. Login as validator
        resp = self.client.post('/api/auth/login', json={'username': 'validator1', 'password': 'password'})
        val_token = json.loads(resp.data)['token']
        
        # 4. Supervisor creates task
        # node_id = 6 (Grading Machine 01), protocol_id = 1 (Daily Alcohol Sanitization), worker_id = 3 (Waleed Worker)
        resp = self.client.post('/api/tasks', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'node_id': 6,
            'protocol_id': 1,
            'worker_id': 3
        })
        self.assertEqual(resp.status_code, 200)
        
        # Get tasks list for worker
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {work_token}'})
        tasks = json.loads(resp.data)
        self.assertEqual(len(tasks), 1)
        task_id = tasks[0]['id']
        self.assertEqual(tasks[0]['status'], 'assigned')
        
        # 5. Worker accepts task
        resp = self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={
            'status': 'accepted'
        })
        self.assertEqual(resp.status_code, 200)
        
        # 6. Worker starts task
        resp = self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={
            'status': 'in_progress'
        })
        self.assertEqual(resp.status_code, 200)
        
        # Check starting inventory of Alcohol (item 1 = 120.0 L)
        db = sqlite3.connect(self.db_path)
        db.row_factory = sqlite3.Row
        item = db.execute("SELECT * FROM inventory WHERE id = 1").fetchone()
        start_stock = item['stock']
        self.assertEqual(start_stock, 120.0)
        
        # 7. Worker submits task (consumes 1.5 L of Alcohol item 1)
        dummy_photo = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
        resp = self.client.post(f'/api/tasks/{task_id}/submit', headers={'Authorization': f'Bearer {work_token}'}, json={
            'notes': 'Completed with no issues',
            'photo_before': dummy_photo,
            'photo_after': dummy_photo,
            'consumptions': [{'item_id': 1, 'quantity': 1.5}]
        })
        self.assertEqual(resp.status_code, 200)
        
        # Check stock deduction
        item = db.execute("SELECT * FROM inventory WHERE id = 1").fetchone()
        self.assertEqual(item['stock'], start_stock - 1.5)
        
        # Verify log entry in inventory_logs
        log = db.execute("SELECT * FROM inventory_logs WHERE task_id = ?", (task_id,)).fetchone()
        self.assertIsNotNone(log)
        self.assertEqual(log['quantity'], -1.5)
        
        # 8. Validator rejects task
        resp = self.client.post(f'/api/tasks/{task_id}/validate', headers={'Authorization': f'Bearer {val_token}'}, json={
            'approved': False,
            'rejection_reason': 'Surfaces still dusty'
        })
        self.assertEqual(resp.status_code, 200)
        
        # Worker sees rejected status
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {work_token}'})
        tasks = json.loads(resp.data)
        self.assertEqual(tasks[0]['status'], 'rejected')
        self.assertEqual(tasks[0]['rejection_reason'], 'Surfaces still dusty')
        
        # 9. Worker restarts and completes task again
        resp = self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={'status': 'accepted'})
        resp = self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={'status': 'in_progress'})
        
        # Submit again, consuming 0.5 L of alcohol
        resp = self.client.post(f'/api/tasks/{task_id}/submit', headers={'Authorization': f'Bearer {work_token}'}, json={
            'notes': 'Re-cleaned thoroughly',
            'photo_before': dummy_photo,
            'photo_after': dummy_photo,
            'consumptions': [{'item_id': 1, 'quantity': 0.5}]
        })
        
        # Check total remaining stock: 120.0 - 1.5 (first run) - 0.5 (second run) = 118.0 L
        item = db.execute("SELECT * FROM inventory WHERE id = 1").fetchone()
        self.assertEqual(item['stock'], 118.0)
        db.close()
        
        # 10. Login as Coordinator to validate task
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        coord_token = json.loads(resp.data)['token']
        
        # 10. Coordinator approves task
        resp = self.client.post(f'/api/tasks/{task_id}/validate', headers={'Authorization': f'Bearer {coord_token}'}, json={
            'approved': True
        })
        self.assertEqual(resp.status_code, 200)
        
        # Confirm completed status
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {work_token}'})
        tasks = json.loads(resp.data)
        self.assertEqual(tasks[0]['status'], 'completed')
        self.assertIsNone(tasks[0]['rejection_reason'])

    def test_supervisor_inventory_crud(self):
        # 1. Login as supervisor
        resp = self.client.post('/api/auth/login', json={'username': 'supervisor1', 'password': 'password'})
        sup_token = json.loads(resp.data)['token']
        
        # 2. Supervisor creates new inventory item
        resp = self.client.post('/api/inventory', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'name': 'Supervisor Soap',
            'category': 'chemical',
            'stock': 100.0,
            'min_stock': 10.0,
            'unit': 'L'
        })
        self.assertEqual(resp.status_code, 200)
        
        # 3. Check item was created
        db = sqlite3.connect(self.db_path)
        db.row_factory = sqlite3.Row
        item = db.execute("SELECT * FROM inventory WHERE name = 'Supervisor Soap'").fetchone()
        self.assertIsNotNone(item)
        item_id = item['id']
        self.assertEqual(item['stock'], 100.0)
        
        # 4. Supervisor updates inventory item
        resp = self.client.put(f'/api/inventory/{item_id}', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'name': 'Supervisor Soap v2',
            'category': 'chemical',
            'min_stock': 15.0,
            'unit': 'L'
        })
        self.assertEqual(resp.status_code, 200)
        
        # 5. Supervisor adjusts stock
        resp = self.client.post(f'/api/inventory/{item_id}/adjust', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'quantity': 50.0,
            'notes': 'Restocked by supervisor'
        })
        self.assertEqual(resp.status_code, 200)
        
        # 6. Verify stock level: 100.0 + 50.0 = 150.0
        item = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,)).fetchone()
        self.assertEqual(item['stock'], 150.0)
        db.close()

    def test_manual_backup(self):
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Trigger backup
        resp = self.client.post('/api/backup', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(resp.status_code, 200)

    def test_coordinator_user_update(self):
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        coord_token = json.loads(resp.data)['token']
        
        # 2. Update worker1's password and name
        # worker1 is user ID 3
        resp = self.client.put('/api/users/3', headers={'Authorization': f'Bearer {coord_token}'}, json={
            'name': 'Waleed Worker Updated',
            'role': 'worker',
            'password': 'newpassword123'
        })
        self.assertEqual(resp.status_code, 200)
        
        # 3. Verify worker1 can log in with new password
        resp = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'newpassword123'
        })
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(data['user']['name'], 'Waleed Worker Updated')

    def test_self_profile_update(self):
        # 1. Login as worker1
        resp = self.client.post('/api/auth/login', json={'username': 'worker1', 'password': 'password'})
        token = json.loads(resp.data)['token']
        
        # 2. Update self details (name, password, dummy base64 avatar)
        dummy_avatar = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
        resp = self.client.put('/api/users/profile', headers={'Authorization': f'Bearer {token}'}, json={
            'name': 'Waleed Self Updated',
            'password': 'selfnewpassword123',
            'profile_image': dummy_avatar
        })
        self.assertEqual(resp.status_code, 200)
        data = json.loads(resp.data)
        self.assertEqual(data['user']['name'], 'Waleed Self Updated')
        self.assertIsNotNone(data['user']['profile_image'])
        avatar_filename = data['user']['profile_image']
        # New format: YYYY/MM/profile_avatar_<token>.png (dated subdirectory)
        self.assertIn('profile_avatar_', avatar_filename)
        self.assertTrue(avatar_filename.endswith('.png') or avatar_filename.endswith('.jpg'))

        # Check avatar file exists in static/uploads/<dated path>
        avatar_path = os.path.join('static/uploads', avatar_filename)
        self.assertTrue(os.path.exists(avatar_path))

        # Cleanup created avatar file
        try:
            os.remove(avatar_path)
        except OSError:
            pass
            
        # 3. Verify worker1 can log in with new password
        resp2 = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'selfnewpassword123'
        })
        self.assertEqual(resp2.status_code, 200)
        data2 = json.loads(resp2.data)
        self.assertEqual(data2['user']['profile_image'], avatar_filename)

    def test_audit_log_filters(self):
        import datetime
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Fetch audit logs filtered by username 'admin'
        resp = self.client.get('/api/audit', headers={'Authorization': f'Bearer {token}'}, query_string={
            'username': 'admin'
        })
        self.assertEqual(resp.status_code, 200)
        logs = json.loads(resp.data)
        self.assertTrue(len(logs) > 0)
        self.assertEqual(logs[0]['username'], 'admin')
        
        # Test with date range filters
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        resp_date = self.client.get('/api/audit', headers={'Authorization': f'Bearer {token}'}, query_string={
            'start_date': today,
            'end_date': today
        })
        self.assertEqual(resp_date.status_code, 200)
        logs_date = json.loads(resp_date.data)
        self.assertTrue(len(logs_date) > 0)

    def test_export_reports_excel(self):
        import io
        import openpyxl
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Call export endpoint using header
        resp_header = self.client.get('/api/reports/export', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(resp_header.status_code, 200)
        self.assertEqual(resp_header.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        # Verify sheets inside the workbook
        wb = openpyxl.load_workbook(io.BytesIO(resp_header.data))
        self.assertEqual(len(wb.sheetnames), 4)
        self.assertIn("Overview", wb.sheetnames)
        self.assertIn("Worker Performance", wb.sheetnames)
        self.assertIn("Machine Performance", wb.sheetnames)
        self.assertIn("Supervisor Activity", wb.sheetnames)

    def test_export_inventory_catalog(self):
        import io
        import openpyxl
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        resp_export = self.client.get('/api/inventory/export', query_string={'token': token})
        self.assertEqual(resp_export.status_code, 200)
        self.assertEqual(resp_export.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        wb = openpyxl.load_workbook(io.BytesIO(resp_export.data))
        self.assertEqual(wb.sheetnames, ["Inventory Catalog"])
        ws = wb.active
        # Check headers
        headers = [ws.cell(row=1, column=col).value for col in range(1, 7)]
        self.assertEqual(headers, ["Product Name", "Category", "Current Stock", "Alert Threshold", "Unit", "Status"])

    def test_export_consumption_logs(self):
        import io
        import openpyxl
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Test export without filters
        resp_export = self.client.get('/api/inventory/logs/export', query_string={'token': token})
        self.assertEqual(resp_export.status_code, 200)
        
        wb = openpyxl.load_workbook(io.BytesIO(resp_export.data))
        self.assertEqual(wb.sheetnames, ["Consumption Logs"])
        ws = wb.active
        headers = [ws.cell(row=1, column=col).value for col in range(1, 8)]
        self.assertEqual(headers, ["Timestamp", "Product Name", "Quantity Consumed", "Unit", "Task ID", "Logged By", "Notes"])
        
        # 3. Test with filters
        resp_filtered = self.client.get('/api/inventory/logs/export', query_string={
            'token': token,
            'username': 'non_existent_user'
        })
        self.assertEqual(resp_filtered.status_code, 200)
        wb_filt = openpyxl.load_workbook(io.BytesIO(resp_filtered.data))
        ws_filt = wb_filt.active
        self.assertEqual(ws_filt.max_row, 1) # Header only, no matching logs

    def test_export_audit_trail(self):
        import io
        import openpyxl
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Test export without filters
        resp_export = self.client.get('/api/audit/export', query_string={'token': token})
        self.assertEqual(resp_export.status_code, 200)
        self.assertEqual(resp_export.mimetype, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        
        wb = openpyxl.load_workbook(io.BytesIO(resp_export.data))
        self.assertEqual(wb.sheetnames, ["Audit Trail"])
        ws = wb.active
        headers = [ws.cell(row=1, column=col).value for col in range(1, 6)]
        self.assertEqual(headers, ["Timestamp", "Actor", "Role", "Action Code", "Details"])
        
        # 3. Test with username filter
        resp_filtered = self.client.get('/api/audit/export', query_string={
            'token': token,
            'username': 'admin'
        })
        self.assertEqual(resp_filtered.status_code, 200)
        wb_filt = openpyxl.load_workbook(io.BytesIO(resp_filtered.data))
        ws_filt = wb_filt.active
        self.assertTrue(ws_filt.max_row > 1)
        # Ensure all rows (except header) have 'admin' as the actor (column 2)
        for row in range(2, ws_filt.max_row + 1):
            self.assertEqual(ws_filt.cell(row=row, column=2).value, 'admin')

    def test_supervisor_validation(self):
        # 1. Login as supervisor
        resp = self.client.post('/api/auth/login', json={'username': 'supervisor1', 'password': 'password'})
        sup_token = json.loads(resp.data)['token']
        
        # 2. Login as worker
        resp = self.client.post('/api/auth/login', json={'username': 'worker1', 'password': 'password'})
        work_token = json.loads(resp.data)['token']
        
        # 3. Supervisor creates task
        resp = self.client.post('/api/tasks', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'node_id': 6,
            'protocol_id': 1,
            'worker_id': 3
        })
        self.assertEqual(resp.status_code, 200)
        
        # Get tasks list
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {work_token}'})
        tasks = json.loads(resp.data)
        task_id = tasks[0]['id']
        
        # 4. Worker accepts, starts and completes task
        self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={'status': 'accepted'})
        self.client.post(f'/api/tasks/{task_id}/status', headers={'Authorization': f'Bearer {work_token}'}, json={'status': 'in_progress'})
        dummy_photo = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAYAAAAfFcSJAAAADUlEQVR42mNk+M9QDwADhgGAWjR9awAAAABJRU5ErkJggg=="
        self.client.post(f'/api/tasks/{task_id}/submit', headers={'Authorization': f'Bearer {work_token}'}, json={
            'notes': 'Done',
            'photo_before': dummy_photo,
            'photo_after': dummy_photo,
            'consumptions': []
        })
        
        # 5. Supervisor validates (rejects) task
        resp = self.client.post(f'/api/tasks/{task_id}/validate', headers={'Authorization': f'Bearer {sup_token}'}, json={
            'approved': False,
            'rejection_reason': 'Please redo'
        })
        self.assertEqual(resp.status_code, 200)
        
        # Verify status is rejected
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {work_token}'})
        tasks = json.loads(resp.data)
        self.assertEqual(tasks[0]['status'], 'rejected')

    def test_account_lockout(self):
        # 1. Fire 4 failed logins - should not lock
        for _ in range(4):
            response = self.client.post('/api/auth/login', json={
                'username': 'worker1',
                'password': 'wrongpassword'
            })
            self.assertEqual(response.status_code, 401)
            
        # 2. Fire 5th failed login - should lock and return 403
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'wrongpassword'
        })
        self.assertEqual(response.status_code, 403)
        self.assertIn('locked', json.loads(response.data)['message'])
        
        # 3. Fire a correct login attempt - should still fail because of lockout (returns 403)
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'password'
        })
        self.assertEqual(response.status_code, 403)
        self.assertIn('locked', json.loads(response.data)['message'])

    def test_session_inactivity_timeout(self):
        import datetime
        # 1. Login to get token
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'password'
        })
        self.assertEqual(response.status_code, 200)
        token = json.loads(response.data)['token']
        
        # 2. Verify token works
        response = self.client.get('/api/inventory', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(response.status_code, 200)
        
        # 3. Manually backdate last_activity by 16 minutes (960 seconds) in ACTIVE_SESSIONS
        from app import ACTIVE_SESSIONS
        ACTIVE_SESSIONS[token]['last_activity'] = datetime.datetime.now() - datetime.timedelta(minutes=16)

        # 4. Verify request now returns 401 due to session timeout
        response = self.client.get('/api/inventory', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(response.status_code, 401)

    def test_session_max_age(self):
        """Sessions older than 8 hours (absolute max) must be rejected regardless of activity."""
        # 1. Login
        response = self.client.post('/api/auth/login', json={
            'username': 'worker1',
            'password': 'password'
        })
        self.assertEqual(response.status_code, 200)
        token = json.loads(response.data)['token']

        # 2. Verify token works
        response = self.client.get('/api/inventory', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(response.status_code, 200)

        # 3. Backdate created_at to 9 hours ago (simulating a stale session)
        ACTIVE_SESSIONS[token]['created_at'] = datetime.datetime.now() - datetime.timedelta(hours=9)
        # Also refresh last_activity so inactivity timeout doesn't fire first
        ACTIVE_SESSIONS[token]['last_activity'] = datetime.datetime.now()

        # 4. Request must now return 401 due to absolute max-age
        response = self.client.get('/api/inventory', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(response.status_code, 401)

    def test_security_headers(self):
        """All API responses must include hardened HTTP security headers."""
        response = self.client.post('/api/auth/login', json={
            'username': 'admin',
            'password': 'admin'
        })
        self.assertEqual(response.status_code, 200)

        # Verify all required security headers are present
        self.assertEqual(response.headers.get('X-Content-Type-Options'), 'nosniff')
        self.assertEqual(response.headers.get('X-Frame-Options'), 'DENY')
        self.assertIn('1', response.headers.get('X-XSS-Protection', ''))
        self.assertIsNotNone(response.headers.get('Content-Security-Policy'))
        self.assertIsNotNone(response.headers.get('Referrer-Policy'))

    def test_upload_security(self):
        """Image upload must reject non-image files and files exceeding 5 MB."""
        from utils.image_utils import save_base64_image, MAX_IMAGE_BYTES

        # --- Test 1: Oversized image (>5 MB) must be rejected ---
        # Create a fake oversized payload (all-zero bytes, >5 MB)
        oversized_data = b'\x89PNG\r\n\x1a\n' + b'\x00' * (MAX_IMAGE_BYTES + 1)
        oversized_b64 = 'data:image/png;base64,' + base64.b64encode(oversized_data).decode()
        result = save_base64_image(oversized_b64, 'test')
        self.assertIsNone(result, 'Oversized image should be rejected')

        # --- Test 2: Non-image file (fake exe) must be rejected ---
        fake_exe = b'MZ\x90\x00' + b'\x00' * 100  # DOS MZ header
        fake_exe_b64 = 'data:application/octet-stream;base64,' + base64.b64encode(fake_exe).decode()
        result = save_base64_image(fake_exe_b64, 'test')
        self.assertIsNone(result, 'Non-image file should be rejected')

        # --- Test 3: Valid small PNG must succeed ---
        # Minimal valid 1x1 PNG
        minimal_png = (
            b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01'
            b'\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f'
            b'\x00\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82'
        )
        valid_b64 = 'data:image/png;base64,' + base64.b64encode(minimal_png).decode()
        result = save_base64_image(valid_b64, 'test')
        self.assertIsNotNone(result, 'Valid PNG should be accepted')
        self.assertTrue(result.endswith('.png'), 'Saved file should have .png extension')

    def test_tasks_enhanced_operations(self):
        """Test task creation with location photo, task editing, and task deleting."""
        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']
        
        # 2. Login as worker
        resp_worker = self.client.post('/api/auth/login', json={'username': 'worker1', 'password': 'password'})
        worker_token = json.loads(resp_worker.data)['token']
        
        # 3. Valid base64 PNG
        minimal_png = (
            b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01'
            b'\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f'
            b'\x00\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82'
        )
        location_photo_b64 = 'data:image/png;base64,' + base64.b64encode(minimal_png).decode()
        
        # Create task with location photo (protocol=1, worker=3, node=6)
        resp = self.client.post('/api/tasks', headers={'Authorization': f'Bearer {token}'}, json={
            'node_id': 6,
            'protocol_id': 1,
            'worker_id': 3,
            'photo_location': location_photo_b64
        })
        self.assertEqual(resp.status_code, 200)
        
        # 4. Worker GET lists task and retrieves hydrated protocol description and required materials
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {worker_token}'})
        tasks = json.loads(resp.data)
        self.assertTrue(len(tasks) >= 1)
        task = next(t for t in tasks if t['worker_id'] == 3)
        self.assertIsNotNone(task['photo_location'])
        self.assertIn('protocol_requirements', task)
        self.assertIn('protocol_duration', task)
        task_id = task['id']
        
        # 5. Coordinator PUT edit task
        resp = self.client.put(f'/api/tasks/{task_id}', headers={'Authorization': f'Bearer {token}'}, json={
            'node_id': 5,
            'protocol_id': 1,
            'worker_id': 3,
            'status': 'accepted',
            'notes': 'Location details edited'
        })
        self.assertEqual(resp.status_code, 200)
        
        # Verify changes
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {worker_token}'})
        tasks = json.loads(resp.data)
        task = next(t for t in tasks if t['id'] == task_id)
        self.assertEqual(task['node_id'], 5)
        self.assertEqual(task['status'], 'accepted')
        self.assertEqual(task['notes'], 'Location details edited')
        
        # 6. Coordinator DELETE task
        resp = self.client.delete(f'/api/tasks/{task_id}', headers={'Authorization': f'Bearer {token}'})
        self.assertEqual(resp.status_code, 200)
        
        # Verify deleted
        resp = self.client.get('/api/tasks', headers={'Authorization': f'Bearer {worker_token}'})
        tasks = json.loads(resp.data)
        task_ids = [t['id'] for t in tasks]
        self.assertNotIn(task_id, task_ids)

    def test_export_tasks(self):
        """Coordinator can export all tasks as an Excel file via /api/tasks/export."""
        import io
        import openpyxl

        # 1. Login as coordinator
        resp = self.client.post('/api/auth/login', json={'username': 'admin', 'password': 'admin'})
        token = json.loads(resp.data)['token']

        # 2. Call tasks export endpoint
        resp_export = self.client.get('/api/tasks/export', query_string={'token': token})
        self.assertEqual(resp_export.status_code, 200)
        self.assertEqual(
            resp_export.mimetype,
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

        # 3. Verify workbook structure
        wb = openpyxl.load_workbook(io.BytesIO(resp_export.data))
        self.assertIn('Sanitation Tasks', wb.sheetnames)
        ws = wb['Sanitation Tasks']
        headers = [ws.cell(row=1, column=col).value for col in range(1, 11)]
        self.assertEqual(headers, [
            'Task ID', 'Protocol', 'Area/Machine', 'Worker',
            'Supervisor', 'Status', 'Start Time', 'End Time',
            'Notes', 'Rejection Reason'
        ])

        # 4. Worker cannot access the export endpoint (forbidden)
        resp_worker = self.client.post('/api/auth/login', json={'username': 'worker1', 'password': 'password'})
        worker_token = json.loads(resp_worker.data)['token']
        resp_denied = self.client.get('/api/tasks/export', query_string={'token': worker_token})
        self.assertEqual(resp_denied.status_code, 403)

if __name__ == '__main__':
    unittest.main()
