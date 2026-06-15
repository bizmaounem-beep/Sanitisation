import os
import sys
import sqlite3
import json
import datetime
import shutil
import threading
import time
import logging
from functools import wraps
from flask import Flask, request, jsonify, render_template, send_from_directory, g, has_app_context
from werkzeug.security import generate_password_hash, check_password_hash
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address

# --- Utility modules ---
from utils.logging_config import configure_logging
from utils.security_utils import generate_secure_token, is_session_expired, make_session
from utils.image_utils import save_base64_image

# --- Configure structured logging (must happen before Flask app init) ---
configure_logging()
logger = logging.getLogger(__name__)

app = Flask(__name__)
DATABASE = 'database.db'

# --- Rate Limiter (in-memory, suitable for factory intranet) ---
limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=[],          # No global limit; applied per-route
    storage_uri='memory://',
    on_breach=lambda limit_info: None  # Handled by Flask-Limiter's 429 response
)

# Active user sessions store (token -> session dict)
# In-memory cache; sessions are also persisted in the DB to survive restarts.
ACTIVE_SESSIONS = {}

# --- Security Headers ---
@app.after_request
def add_security_headers(response):
    """Inject hardened HTTP security headers on every response."""
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    response.headers['Referrer-Policy'] = 'strict-origin'
    response.headers['Content-Security-Policy'] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com https://fonts.gstatic.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data: blob:;"
    )
    return response

# Ensure static directories exist
os.makedirs('static/uploads', exist_ok=True)
os.makedirs('backups', exist_ok=True)
os.makedirs('logs', exist_ok=True)

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DATABASE)
        db.row_factory = sqlite3.Row
        # Enable Foreign Keys
        db.execute("PRAGMA foreign_keys = ON;")
    return db

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    """Initializes the database using schema.sql and auto-hashes seed passwords."""
    if not os.path.exists(DATABASE):
        with app.app_context():
            db = get_db()
            with app.open_resource('schema.sql', mode='r', encoding='utf-8') as f:
                db.cursor().executescript(f.read())
            
            # Hash seed passwords
            cursor = db.cursor()
            cursor.execute("SELECT id, password FROM users")
            users = cursor.fetchall()
            for u in users:
                user_id = u['id']
                raw_pwd = u['password']
                if not (raw_pwd.startswith('pbkdf2:') or raw_pwd.startswith('scrypt:')):
                    hashed = generate_password_hash(raw_pwd)
                    cursor.execute("UPDATE users SET password = ? WHERE id = ?", (hashed, user_id))
            db.commit()
            print("Database initialized and passwords auto-hashed.")
    else:
        # Run automatic migration if profile_image column doesn't exist
        with app.app_context():
            db = get_db()
            cursor = db.cursor()
            cursor.execute("PRAGMA table_info(users)")
            columns = [col[1] for col in cursor.fetchall()]
            if 'profile_image' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN profile_image TEXT")
                    db.commit()
                    print("Migration: Added profile_image column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)
            if 'reset_code' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN reset_code TEXT")
                    db.commit()
                    print("Migration: Added reset_code column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)
            if 'reset_code_expires' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN reset_code_expires DATETIME")
                    db.commit()
                    print("Migration: Added reset_code_expires column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)
            if 'reset_requested' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN reset_requested INTEGER DEFAULT 0")
                    db.commit()
                    print("Migration: Added reset_requested column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)
            if 'failed_login_attempts' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN failed_login_attempts INTEGER DEFAULT 0")
                    db.commit()
                    print("Migration: Added failed_login_attempts column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)
            if 'locked_until' not in columns:
                try:
                    cursor.execute("ALTER TABLE users ADD COLUMN locked_until TEXT")
                    db.commit()
                    print("Migration: Added locked_until column to users table.")
                except Exception as e:
                    print(f"Migration error: {e}", file=sys.stderr)

            # Auto-migrate tasks table to add photo_location if missing
            cursor.execute("PRAGMA table_info(tasks)")
            task_columns = [col[1] for col in cursor.fetchall()]
            if 'photo_location' not in task_columns:
                try:
                    cursor.execute("ALTER TABLE tasks ADD COLUMN photo_location TEXT")
                    db.commit()
                    print("Migration: Added photo_location column to tasks table.")
                except Exception as e:
                    print(f"Migration error (tasks): {e}", file=sys.stderr)

        # --- Persist-sessions migration: create sessions table if missing ---
        with app.app_context():
            db = get_db()
            db.execute("""
                CREATE TABLE IF NOT EXISTS sessions (
                    token TEXT PRIMARY KEY,
                    user_json TEXT NOT NULL,
                    last_activity TEXT NOT NULL,
                    created_at TEXT NOT NULL
                )
            """)
            db.commit()
            # Load existing valid sessions into in-memory cache
            now = datetime.datetime.now()
            rows = db.execute("SELECT * FROM sessions").fetchall()
            for row in rows:
                try:
                    last_act = datetime.datetime.strptime(row['last_activity'], '%Y-%m-%d %H:%M:%S')
                    created = datetime.datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S')
                    sess = {
                        'user': json.loads(row['user_json']),
                        'last_activity': last_act,
                        'created_at': created
                    }
                    if not is_session_expired(sess):
                        ACTIVE_SESSIONS[row['token']] = sess
                except Exception:
                    pass
            # Purge expired sessions from DB using timezone-independent comparison
            now = datetime.datetime.now()
            inactivity_cutoff = (now - datetime.timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
            created_cutoff = (now - datetime.timedelta(hours=8)).strftime('%Y-%m-%d %H:%M:%S')
            db.execute("""
                DELETE FROM sessions WHERE
                    last_activity < ? OR created_at < ?
            """, (inactivity_cutoff, created_cutoff))
            db.commit()
            print(f"Sessions restored: {len(ACTIVE_SESSIONS)} active session(s) loaded from DB.")

# --- Authentication Helpers ---
def get_current_user():
    auth_header = request.headers.get('Authorization', None)
    token = None
    if auth_header and auth_header.startswith('Bearer '):
        token = auth_header.split(' ')[1]
    if not token:
        token = request.args.get('token')

    if not token:
        return None

    if token not in ACTIVE_SESSIONS:
        # Try to restore session from DB
        try:
            db = get_db()
            row = db.execute("SELECT * FROM sessions WHERE token = ?", (token,)).fetchone()
            if row:
                last_act = datetime.datetime.strptime(row['last_activity'], '%Y-%m-%d %H:%M:%S')
                created = datetime.datetime.strptime(row['created_at'], '%Y-%m-%d %H:%M:%S')
                sess = {
                    'user': json.loads(row['user_json']),
                    'last_activity': last_act,
                    'created_at': created
                }
                if not is_session_expired(sess):
                    ACTIVE_SESSIONS[token] = sess
        except Exception as e:
            logger.error('Error restoring session from DB: %s', str(e))

    if token in ACTIVE_SESSIONS:
        session = ACTIVE_SESSIONS[token]
        # Check both inactivity timeout (15 min) and absolute max-age (8 hours)
        if is_session_expired(session):
            logger.info('Session expired for token (inactivity or shift end), removing.')
            del ACTIVE_SESSIONS[token]
            # Remove from persistent DB
            try:
                db = get_db()
                db.execute("DELETE FROM sessions WHERE token = ?", (token,))
                db.commit()
            except Exception:
                pass
            return None
        now = datetime.datetime.now()
        session['last_activity'] = now
        # Persist updated last_activity to DB so it survives restarts
        try:
            db = get_db()
            db.execute(
                "UPDATE sessions SET last_activity = ? WHERE token = ?",
                (now.strftime('%Y-%m-%d %H:%M:%S'), token)
            )
            db.commit()
        except Exception:
            pass
        return session['user']
    return None


def login_required(roles=None):
    def decorator(f):
        @wraps(f)
        def decorated_function(*args, **kwargs):
            user = get_current_user()
            if not user:
                return jsonify({'message': 'Unauthorized. Please log in.'}), 401
            if roles and user['role'] not in roles:
                return jsonify({'message': 'Forbidden. Insufficient permissions.'}), 403
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def log_audit(user_id, username, role, action, details=None):
    """Logs system events to the audit trail table using active request database connection or fallback."""
    if has_app_context():
        try:
            db = get_db()
            db.execute(
                "INSERT INTO audit_logs (user_id, username, role, action, details) VALUES (?, ?, ?, ?, ?)",
                (user_id, username, role, action, details)
            )
        except Exception as e:
            logger.error('Audit Log Error (in request context): %s', str(e))
    else:
        try:
            db = sqlite3.connect(DATABASE)
            db.execute(
                "INSERT INTO audit_logs (user_id, username, role, action, details) VALUES (?, ?, ?, ?, ?)",
                (user_id, username, role, action, details)
            )
            db.commit()
            db.close()
        except Exception as e:
            logger.error('Audit Log Error (no request context): %s', str(e))

# --- Automatic Backup System ---
def backup_db():
    """Performs SQLite database backup, keeping only the last 10 backups."""
    try:
        timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_file = f"backups/database_{timestamp}.db"
        shutil.copyfile(DATABASE, backup_file)

        # Remove old backups if exceeding 10
        backups = sorted([f for f in os.listdir('backups') if f.endswith('.db')])
        while len(backups) > 10:
            os.remove(os.path.join('backups', backups.pop(0)))
        logger.info('Automatic backup created: %s', backup_file)
    except Exception as e:
        logger.error('Backup Error: %s', str(e))

def backup_scheduler():
    """Runs in a background thread to copy the database file every 24 hours."""
    while True:
        time.sleep(24 * 3600)  # Wait 24 hours
        backup_db()

# Start background backup scheduler thread
threading.Thread(target=backup_scheduler, daemon=True).start()

# --- Page Routes ---
@app.route('/')
def index_page():
    return render_template('index.html')

@app.route('/static/uploads/<path:filename>')
def serve_uploads(filename):
    return send_from_directory('static/uploads', filename)

# --- Auth API ---
@app.route('/api/auth/login', methods=['POST'])
@limiter.limit('5 per minute')
def auth_login():
    data = request.json or {}
    username = data.get('username')
    password = data.get('password')

    if not username or not password:
        return jsonify({'message': 'Missing username or password'}), 400

    db = get_db()
    user = db.execute("SELECT * FROM users WHERE username = ? AND is_active = 1", (username,)).fetchone()

    if user and user['locked_until']:
        try:
            locked_until_dt = datetime.datetime.strptime(user['locked_until'], '%Y-%m-%d %H:%M:%S')
            if datetime.datetime.now() < locked_until_dt:
                logger.warning('Blocked login attempt for locked account: %s', username)
                return jsonify({'message': 'Account temporarily locked due to too many failed login attempts. Please try again later.'}), 403
        except Exception:
            pass

    if user and check_password_hash(user['password'], password):
        if not user['is_approved']:
            logger.warning('Login attempt for unapproved account: %s', username)
            return jsonify({'message': 'Account pending coordinator approval'}), 403

        # Success: reset attempts
        db.execute("UPDATE users SET failed_login_attempts = 0, locked_until = NULL WHERE id = ?", (user['id'],))
        db.commit()

        token = generate_secure_token()
        user_data = {
            'id': user['id'],
            'username': user['username'],
            'name': user['name'],
            'role': user['role'],
            'language': user['language'],
            'profile_image': user['profile_image']
        }
        sess = make_session(user_data)
        ACTIVE_SESSIONS[token] = sess
        # Persist session to DB so it survives server restarts
        now_str = sess['created_at'].strftime('%Y-%m-%d %H:%M:%S')
        db.execute(
            "INSERT OR REPLACE INTO sessions (token, user_json, last_activity, created_at) VALUES (?, ?, ?, ?)",
            (token, json.dumps(user_data), now_str, now_str)
        )
        logger.info('Successful login: %s (%s)', username, user['role'])
        log_audit(user['id'], user['username'], user['role'], 'LOGIN', 'Standard user login')
        db.commit()
        return jsonify({'user': user_data, 'token': token})


    if user:
        attempts = (user['failed_login_attempts'] or 0) + 1
        if attempts >= 5:
            locked_time = (datetime.datetime.now() + datetime.timedelta(minutes=15)).strftime('%Y-%m-%d %H:%M:%S')
            db.execute("UPDATE users SET failed_login_attempts = ?, locked_until = ? WHERE id = ?", (attempts, locked_time, user['id']))
            db.commit()
            logger.warning('Account locked after 5 failed attempts: %s', username)
            return jsonify({'message': 'Account temporarily locked due to 5 failed login attempts. Please try again in 15 minutes.'}), 403
        else:
            db.execute("UPDATE users SET failed_login_attempts = ? WHERE id = ?", (attempts, user['id']))
            db.commit()
            logger.warning('Failed login attempt %d/5 for user: %s', attempts, username)

    return jsonify({'message': 'Invalid username or password'}), 401


@app.route('/api/auth/register', methods=['POST'])
@limiter.limit('10 per hour')
def auth_register():
    data = request.json or {}
    username = data.get('username')
    password = data.get('password')
    name = data.get('name')
    role = data.get('role')
    personal_code = data.get('personal_code')
    
    if not username or not password or not name or not role:
        return jsonify({'message': 'Missing required registration fields'}), 400
        
    db = get_db()
    existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return jsonify({'message': 'Username already exists'}), 400
        
    if personal_code:
        existing_code = db.execute("SELECT id FROM users WHERE personal_code = ?", (personal_code,)).fetchone()
        if existing_code:
            return jsonify({'message': 'Personal code already in use'}), 400
            
    is_approved = 0 if role == 'supervisor' else 1
    hashed_pwd = generate_password_hash(password)
    
    cursor = db.cursor()
    cursor.execute(
        "INSERT INTO users (username, password, name, role, personal_code, is_approved) VALUES (?, ?, ?, ?, ?, ?)",
        (username, hashed_pwd, name, role, personal_code or None, is_approved)
    )
    new_user_id = cursor.lastrowid
    log_audit(new_user_id, username, role, 'REGISTER', f"Registered as {role} (approved: {is_approved})")
    db.commit()
    return jsonify({'message': 'Registration successful', 'pending_approval': not is_approved})


@app.route('/api/auth/request_reset', methods=['POST'])
@limiter.limit('5 per hour')
def auth_request_reset():
    data = request.json or {}
    username = data.get('username')
    name = data.get('name')
    
    if not username or not name:
        return jsonify({'message': 'Missing username or name'}), 400
        
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND LOWER(name) = LOWER(?) AND is_active = 1",
        (username.strip(), name.strip())
    ).fetchone()
    
    if not user:
        return jsonify({'message': 'Verification failed. Username and Full Name do not match our records.'}), 404
        
    db.execute("UPDATE users SET reset_requested = 1 WHERE id = ?", (user['id'],))
    log_audit(user['id'], user['username'], user['role'], 'REQUEST_RESET', 'Requested password/PIN reset from Coordinator')
    db.commit()
    return jsonify({'message': 'Reset request submitted successfully. Contact coordinator for the reset code.'})


@app.route('/api/auth/recover', methods=['POST'])
def auth_recover():
    data = request.json or {}
    username = data.get('username')
    name = data.get('name')
    reset_code = data.get('reset_code')
    
    if not username or not name or not reset_code:
        return jsonify({'message': 'Missing username, name, or reset code'}), 400
        
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND LOWER(name) = LOWER(?) AND reset_code = ? AND is_active = 1",
        (username.strip(), name.strip(), reset_code.strip())
    ).fetchone()
    
    if not user:
        return jsonify({'message': 'Verification failed. Username, Full Name, or Reset Code do not match.'}), 404
        
    if user['reset_code_expires']:
        expires_dt = datetime.datetime.strptime(user['reset_code_expires'], '%Y-%m-%d %H:%M:%S')
        if datetime.datetime.now() > expires_dt:
            return jsonify({'message': 'Reset code has expired.'}), 400
            
    return jsonify({
        'status': 'success',
        'role': user['role'],
        'requires_reset': True
    })


@app.route('/api/auth/reset_password', methods=['POST'])
def auth_reset_password():
    data = request.json or {}
    username = data.get('username')
    name = data.get('name')
    reset_code = data.get('reset_code')
    new_password = data.get('new_password')
    
    if not username or not name or not reset_code or not new_password:
        return jsonify({'message': 'Missing required fields'}), 400
        
    db = get_db()
    user = db.execute(
        "SELECT * FROM users WHERE LOWER(username) = LOWER(?) AND LOWER(name) = LOWER(?) AND reset_code = ? AND is_active = 1",
        (username.strip(), name.strip(), reset_code.strip())
    ).fetchone()
    
    if not user:
        return jsonify({'message': 'Verification failed. User or reset code not found.'}), 404
        
    if user['reset_code_expires']:
        expires_dt = datetime.datetime.strptime(user['reset_code_expires'], '%Y-%m-%d %H:%M:%S')
        if datetime.datetime.now() > expires_dt:
            return jsonify({'message': 'Reset code has expired.'}), 400
            
    hashed_pwd = generate_password_hash(new_password)
    db.execute("UPDATE users SET password = ?, reset_code = NULL, reset_code_expires = NULL WHERE id = ?", 
               (hashed_pwd, user['id']))
    log_audit(user['id'], user['username'], user['role'], 'RESET_PASSWORD', 'Password reset via coordinator reset code')
    db.commit()
    return jsonify({'message': 'Password reset successfully!'})


# --- User Management API (Coordinator Only) ---
@app.route('/api/users', methods=['GET', 'POST'])
@login_required(['coordinator'])
def users_list_create():
    db = get_db()
    current_user = get_current_user()
    
    if request.method == 'GET':
        rows = db.execute("SELECT id, username, name, role, personal_code, is_approved, is_active, profile_image, reset_requested, created_at FROM users").fetchall()
        return jsonify([dict(r) for r in rows])
        
    # POST
    data = request.json or {}
    username = data.get('username')
    password = data.get('password')
    name = data.get('name')
    role = data.get('role')
    personal_code = data.get('personal_code')
    
    if not username or not password or not name or not role:
        return jsonify({'message': 'Missing required user fields'}), 400
        
    existing = db.execute("SELECT id FROM users WHERE username = ?", (username,)).fetchone()
    if existing:
        return jsonify({'message': 'Username already exists'}), 400
        
    hashed_pwd = generate_password_hash(password)
    db.execute(
        "INSERT INTO users (username, password, name, role, personal_code, is_approved) VALUES (?, ?, ?, ?, ?, 1)",
        (username, hashed_pwd, name, role, personal_code or None)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'CREATE_USER', f"Created user {username} ({role})")
    db.commit()
    return jsonify({'message': 'User created successfully'})

@app.route('/api/users/<int:user_id>', methods=['PUT', 'DELETE'])
@login_required(['coordinator'])
def users_update_delete(user_id):
    db = get_db()
    current_user = get_current_user()
    
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'message': 'User not found'}), 404
        
    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', user['name'])
        role = data.get('role', user['role'])
        personal_code = data.get('personal_code', user['personal_code'])
        is_active = data.get('is_active', user['is_active'])
        password = data.get('password')
        profile_image_b64 = data.get('profile_image')
        
        profile_image_fn = user['profile_image']
        if profile_image_b64:
            profile_image_fn = save_base64_image(profile_image_b64, 'profile_avatar')
            
        if password:
            hashed_pwd = generate_password_hash(password)
            db.execute("UPDATE users SET name=?, role=?, personal_code=?, is_active=?, password=?, profile_image=? WHERE id=?", 
                       (name, role, personal_code or None, is_active, hashed_pwd, profile_image_fn, user_id))
        else:
            db.execute("UPDATE users SET name=?, role=?, personal_code=?, is_active=?, profile_image=? WHERE id=?", 
                       (name, role, personal_code or None, is_active, profile_image_fn, user_id))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'UPDATE_USER', f"Updated user {user['username']}")
        db.commit()
        return jsonify({'message': 'User updated successfully'})
        
    # DELETE (Soft Deactivate)
    db.execute("UPDATE users SET is_active = 0 WHERE id = ?", (user_id,))
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'DEACTIVATE_USER', f"Deactivated user {user['username']}")
    db.commit()
    return jsonify({'message': 'User deactivated successfully'})


@app.route('/api/users/<int:user_id>/reset_code', methods=['POST'])
@login_required(['coordinator'])
def users_generate_reset_code(user_id):
    import random
    db = get_db()
    current_user = get_current_user()
    
    user = db.execute("SELECT * FROM users WHERE id = ? AND is_active = 1", (user_id,)).fetchone()
    if not user:
        return jsonify({'message': 'User not found or inactive'}), 404
        
    code = f"{random.randint(100000, 999999)}"
    expires_at = (datetime.datetime.now() + datetime.timedelta(hours=24)).strftime('%Y-%m-%d %H:%M:%S')
    
    db.execute(
        "UPDATE users SET reset_code = ?, reset_code_expires = ?, reset_requested = 0 WHERE id = ?",
        (code, expires_at, user_id)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'GENERATE_RESET_CODE', f"Generated reset code for {user['username']}")
    db.commit()
    return jsonify({'reset_code': code})


@app.route('/api/users/approve/<int:user_id>', methods=['POST'])
@login_required(['coordinator'])
def users_approve(user_id):
    db = get_db()
    current_user = get_current_user()
    
    data = request.json or {}
    approve = data.get('approve', True)
    
    user = db.execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        return jsonify({'message': 'User not found'}), 404
        
    if approve:
        db.execute("UPDATE users SET is_approved = 1 WHERE id = ?", (user_id,))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'APPROVE_USER', f"Approved user {user['username']}")
        db.commit()
        return jsonify({'message': 'User approved successfully'})
    else:
        db.execute("DELETE FROM users WHERE id = ?", (user_id,))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'REJECT_USER', f"Rejected registration for {user['username']}")
        db.commit()
        return jsonify({'message': 'User registration rejected and removed'})

@app.route('/api/users/profile', methods=['PUT'])
@login_required()
def update_profile():
    user = get_current_user()
    if not user:
        return jsonify({'message': 'Unauthorized'}), 401
        
    data = request.json or {}
    name = data.get('name')
    password = data.get('password')
    profile_image_b64 = data.get('profile_image')
    
    db = get_db()
    db_user = db.execute("SELECT * FROM users WHERE id = ?", (user['id'],)).fetchone()
    if not db_user:
        return jsonify({'message': 'User not found'}), 404
        
    updated_name = name if name else db_user['name']
    updated_profile_image = db_user['profile_image']
    
    if profile_image_b64:
        updated_profile_image = save_base64_image(profile_image_b64, 'profile_avatar')
        
    if password:
        hashed_pwd = generate_password_hash(password)
        db.execute("UPDATE users SET name = ?, password = ?, profile_image = ? WHERE id = ?",
                   (updated_name, hashed_pwd, updated_profile_image, user['id']))
    else:
        db.execute("UPDATE users SET name = ?, profile_image = ? WHERE id = ?",
                   (updated_name, updated_profile_image, user['id']))
                   
    db.commit()
    
    # Update active sessions
    updated_user_data = {
        'id': user['id'],
        'username': user['username'],
        'name': updated_name,
        'role': user['role'],
        'language': user['language'],
        'profile_image': updated_profile_image
    }
    
    for token, session in ACTIVE_SESSIONS.items():
        if session['user']['id'] == user['id']:
            session['user'] = updated_user_data
            
    log_audit(user['id'], user['username'], user['role'], 'UPDATE_PROFILE', f"Updated their own profile details")
    
    return jsonify({'message': 'Profile updated successfully', 'user': updated_user_data})

@app.route('/api/users/profile/language', methods=['POST'])
@login_required()
def update_profile_language():
    user = get_current_user()
    if not user:
        return jsonify({'message': 'Unauthorized'}), 401
    lang = request.json.get('language')
    if lang not in ['en', 'fr', 'ar']:
        return jsonify({'message': 'Invalid language choice'}), 400
    db = get_db()
    db.execute("UPDATE users SET language = ? WHERE id = ?", (lang, user['id']))
    
    # Update current session dictionary if active
    for token, session in ACTIVE_SESSIONS.items():
        if session['user']['id'] == user['id']:
            session['user']['language'] = lang
            
    db.commit()
    return jsonify({'message': 'Language updated successfully'})

# --- Facility Hierarchy API ---
@app.route('/api/facility', methods=['GET', 'POST'])
@login_required(['coordinator', 'supervisor'])
def facility_list_create():
    db = get_db()
    current_user = get_current_user()
    
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM facility_nodes WHERE is_active = 1").fetchall()
        return jsonify([dict(r) for r in rows])
        
    # POST (Coordinator only can create/modify)
    if current_user['role'] != 'coordinator':
        return jsonify({'message': 'Forbidden'}), 403
        
    data = request.json or {}
    parent_id = data.get('parent_id')
    type = data.get('type')
    name = data.get('name')
    description = data.get('description')
    assigned_protocol_id = data.get('assigned_protocol_id')
    
    if not type or not name:
        return jsonify({'message': 'Missing required fields'}), 400
        
    db.execute(
        "INSERT INTO facility_nodes (parent_id, type, name, description, assigned_protocol_id) VALUES (?, ?, ?, ?, ?)",
        (parent_id or None, type, name, description, assigned_protocol_id or None)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'CREATE_FACILITY_NODE', f"Created node {name} ({type})")
    db.commit()
    return jsonify({'message': 'Facility node created successfully'})

@app.route('/api/facility/<int:node_id>', methods=['PUT', 'DELETE'])
@login_required(['coordinator'])
def facility_update_delete(node_id):
    db = get_db()
    current_user = get_current_user()
    
    node = db.execute("SELECT * FROM facility_nodes WHERE id = ?", (node_id,)).fetchone()
    if not node:
        return jsonify({'message': 'Node not found'}), 404
        
    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', node['name'])
        description = data.get('description', node['description'])
        status = data.get('status', node['status'])
        assigned_protocol_id = data.get('assigned_protocol_id')
        
        actual_protocol = assigned_protocol_id if assigned_protocol_id != "" else None
        
        db.execute(
            "UPDATE facility_nodes SET name = ?, description = ?, status = ?, assigned_protocol_id = ? WHERE id = ?",
            (name, description, status, actual_protocol, node_id)
        )
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'UPDATE_FACILITY_NODE', f"Updated node {name}")
        db.commit()
        return jsonify({'message': 'Facility node updated successfully'})
        
    # DELETE (Soft delete)
    db.execute("UPDATE facility_nodes SET is_active = 0 WHERE id = ?", (node_id,))
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'DELETE_FACILITY_NODE', f"Soft deleted node {node['name']}")
    db.commit()
    return jsonify({'message': 'Facility node removed successfully'})

# --- Sanitation Protocol Management API ---
@app.route('/api/protocols', methods=['GET', 'POST'])
@login_required(['coordinator', 'supervisor'])
def protocols_list_create():
    db = get_db()
    current_user = get_current_user()
    
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM protocols WHERE is_active = 1").fetchall()
        protocols_list = []
        for r in rows:
            p_dict = dict(r)
            p_dict['steps'] = json.loads(p_dict['steps'])
            
            reqs = db.execute(
                "SELECT pr.quantity_required, i.id as item_id, i.name as item_name, i.unit "
                "FROM protocol_requirements pr JOIN inventory i ON pr.item_id = i.id "
                "WHERE pr.protocol_id = ?", (p_dict['id'],)
            ).fetchall()
            p_dict['requirements'] = [dict(req) for req in reqs]
            protocols_list.append(p_dict)
        return jsonify(protocols_list)
        
    # POST (Coordinator or Supervisor)
    data = request.json or {}
    name = data.get('name')
    description = data.get('description')
    steps = data.get('steps', [])
    estimated_duration = data.get('estimated_duration')
    requirements = data.get('requirements', [])
    
    if not name or not steps or not estimated_duration:
        return jsonify({'message': 'Missing required protocol fields'}), 400
        
    cursor = db.cursor()
    cursor.execute(
        "INSERT INTO protocols (name, description, steps, estimated_duration) VALUES (?, ?, ?, ?)",
        (name, description, json.dumps(steps), estimated_duration)
    )
    protocol_id = cursor.lastrowid
    
    for r in requirements:
        cursor.execute(
            "INSERT INTO protocol_requirements (protocol_id, item_id, quantity_required) VALUES (?, ?, ?)",
            (protocol_id, r['item_id'], r['quantity_required'])
        )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'CREATE_PROTOCOL', f"Created protocol {name}")
    db.commit()
    return jsonify({'message': 'Sanitation protocol created successfully'})

@app.route('/api/protocols/<int:protocol_id>', methods=['PUT', 'DELETE'])
@login_required(['coordinator', 'supervisor'])
def protocols_update_delete(protocol_id):
    db = get_db()
    current_user = get_current_user()
    
    proto = db.execute("SELECT * FROM protocols WHERE id = ?", (protocol_id,)).fetchone()
    if not proto:
        return jsonify({'message': 'Protocol not found'}), 404
        
    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', proto['name'])
        description = data.get('description', proto['description'])
        steps = data.get('steps', json.loads(proto['steps']))
        estimated_duration = data.get('estimated_duration', proto['estimated_duration'])
        requirements = data.get('requirements', [])
        
        db.execute(
            "UPDATE protocols SET name = ?, description = ?, steps = ?, estimated_duration = ? WHERE id = ?",
            (name, description, json.dumps(steps), estimated_duration, protocol_id)
        )
        
        db.execute("DELETE FROM protocol_requirements WHERE protocol_id = ?", (protocol_id,))
        for r in requirements:
            db.execute(
                "INSERT INTO protocol_requirements (protocol_id, item_id, quantity_required) VALUES (?, ?, ?)",
                (protocol_id, r['item_id'], r['quantity_required'])
            )
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'UPDATE_PROTOCOL', f"Updated protocol {name}")
        db.commit()
        return jsonify({'message': 'Protocol updated successfully'})
        
    # DELETE (Soft delete)
    db.execute("UPDATE protocols SET is_active = 0 WHERE id = ?", (protocol_id,))
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'DELETE_PROTOCOL', f"Deleted protocol {proto['name']}")
    db.commit()
    return jsonify({'message': 'Protocol removed successfully'})

# --- Inventory Management API ---
@app.route('/api/inventory', methods=['GET', 'POST'])
@login_required(['coordinator', 'supervisor', 'worker'])
def inventory_list_create():
    db = get_db()
    current_user = get_current_user()
    
    if request.method == 'GET':
        rows = db.execute("SELECT * FROM inventory WHERE is_active = 1").fetchall()
        return jsonify([dict(r) for r in rows])
        
    # POST (Coordinator or Supervisor can create stock templates)
    if current_user['role'] not in ['coordinator', 'supervisor']:
        return jsonify({'message': 'Forbidden'}), 403
        
    data = request.json or {}
    name = data.get('name')
    category = data.get('category')
    stock = data.get('stock', 0.0)
    min_stock = data.get('min_stock', 0.0)
    unit = data.get('unit')
    
    if not name or not category or not unit:
        return jsonify({'message': 'Missing required fields'}), 400
        
    db.execute(
        "INSERT INTO inventory (name, category, stock, min_stock, unit) VALUES (?, ?, ?, ?, ?)",
        (name, category, stock, min_stock, unit)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'CREATE_INVENTORY_ITEM', f"Created inventory item {name}")
    db.commit()
    return jsonify({'message': 'Inventory item added successfully'})

@app.route('/api/inventory/<int:item_id>', methods=['PUT', 'DELETE'])
@login_required(['coordinator', 'supervisor'])
def inventory_update_delete(item_id):
    db = get_db()
    current_user = get_current_user()
    
    item = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,)).fetchone()
    if not item:
        return jsonify({'message': 'Inventory item not found'}), 404
        
    if request.method == 'PUT':
        data = request.json or {}
        name = data.get('name', item['name'])
        category = data.get('category', item['category'])
        min_stock = data.get('min_stock', item['min_stock'])
        unit = data.get('unit', item['unit'])
        
        db.execute(
            "UPDATE inventory SET name = ?, category = ?, min_stock = ?, unit = ? WHERE id = ?",
            (name, category, min_stock, unit, item_id)
        )
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'UPDATE_INVENTORY_ITEM', f"Updated item {name}")
        db.commit()
        return jsonify({'message': 'Inventory item updated successfully'})
        
    # DELETE (Soft delete)
    db.execute("UPDATE inventory SET is_active = 0 WHERE id = ?", (item_id,))
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'DELETE_INVENTORY_ITEM', f"Deleted item {item['name']}")
    db.commit()
    return jsonify({'message': 'Inventory item removed successfully'})

@app.route('/api/inventory/<int:item_id>/adjust', methods=['POST'])
@login_required(['coordinator', 'supervisor'])
def inventory_adjust_stock(item_id):
    db = get_db()
    current_user = get_current_user()
    
    item = db.execute("SELECT * FROM inventory WHERE id = ?", (item_id,)).fetchone()
    if not item:
        return jsonify({'message': 'Inventory item not found'}), 404
        
    data = request.json or {}
    quantity = data.get('quantity')
    notes = data.get('notes', '')
    
    if quantity is None:
        return jsonify({'message': 'Quantity is required'}), 400
        
    qty = float(quantity)
    db.execute("UPDATE inventory SET stock = stock + ? WHERE id = ?", (qty, item_id))
    db.execute(
        "INSERT INTO inventory_logs (item_id, quantity, user_id, notes) VALUES (?, ?, ?, ?)",
        (item_id, qty, current_user['id'], notes or 'Manual adjustment')
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'ADJUST_STOCK', f"Adjusted {item['name']} by {qty}")
    db.commit()
    return jsonify({'message': 'Stock adjusted successfully'})

@app.route('/api/inventory/logs', methods=['GET'])
@login_required(['coordinator', 'supervisor'])
def inventory_logs():
    db = get_db()
    item_id = request.args.get('item_id')
    if item_id:
        rows = db.execute(
            "SELECT il.*, i.name as item_name, i.unit, u.name as user_name "
            "FROM inventory_logs il "
            "JOIN inventory i ON il.item_id = i.id "
            "LEFT JOIN users u ON il.user_id = u.id "
            "WHERE il.item_id = ? "
            "ORDER BY il.created_at DESC", (item_id,)
        ).fetchall()
    else:
        rows = db.execute(
            "SELECT il.*, i.name as item_name, i.unit, u.name as user_name "
            "FROM inventory_logs il "
            "JOIN inventory i ON il.item_id = i.id "
            "LEFT JOIN users u ON il.user_id = u.id "
            "ORDER BY il.created_at DESC"
        ).fetchall()
    return jsonify([dict(r) for r in rows])

# --- Tasks Workflow API ---
@app.route('/api/tasks', methods=['GET', 'POST'])
@login_required(['coordinator', 'supervisor', 'worker', 'validator'])
def tasks_list_create():
    db = get_db()
    current_user = get_current_user()
    
    if request.method == 'GET':
        if current_user['role'] == 'worker':
            query = "SELECT t.*, p.name as protocol_name, p.steps as protocol_steps, p.description as protocol_description, p.estimated_duration as protocol_duration, fn.name as node_name " \
                    "FROM tasks t JOIN protocols p ON t.protocol_id = p.id JOIN facility_nodes fn ON t.node_id = fn.id " \
                    "WHERE t.worker_id = ? ORDER BY t.id DESC"
            rows = db.execute(query, (current_user['id'],)).fetchall()
        elif current_user['role'] == 'validator':
            query = "SELECT t.*, p.name as protocol_name, p.steps as protocol_steps, p.description as protocol_description, p.estimated_duration as protocol_duration, fn.name as node_name, w.name as worker_name, s.name as supervisor_name " \
                    "FROM tasks t JOIN protocols p ON t.protocol_id = p.id JOIN facility_nodes fn ON t.node_id = fn.id " \
                    "JOIN users w ON t.worker_id = w.id JOIN users s ON t.supervisor_id = s.id " \
                    "WHERE t.status = 'pending_validation' ORDER BY t.id DESC"
            rows = db.execute(query).fetchall()
        else:
            query = "SELECT t.*, p.name as protocol_name, p.steps as protocol_steps, p.description as protocol_description, p.estimated_duration as protocol_duration, fn.name as node_name, w.name as worker_name, s.name as supervisor_name " \
                    "FROM tasks t JOIN protocols p ON t.protocol_id = p.id JOIN facility_nodes fn ON t.node_id = fn.id " \
                    "JOIN users w ON t.worker_id = w.id JOIN users s ON t.supervisor_id = s.id " \
                    "ORDER BY t.id DESC"
            rows = db.execute(query).fetchall()
            
        result = []
        for r in rows:
            t_dict = dict(r)
            if 'protocol_steps' in t_dict and t_dict['protocol_steps']:
                t_dict['protocol_steps'] = json.loads(t_dict['protocol_steps'])
            
            # Actual consumptions
            cons = db.execute(
                "SELECT tc.quantity_used, i.name as item_name, i.unit "
                "FROM task_consumptions tc JOIN inventory i ON tc.item_id = i.id "
                "WHERE tc.task_id = ?", (t_dict['id'],)
            ).fetchall()
            t_dict['consumptions'] = [dict(c) for c in cons]
            
            # Protocol requirements
            reqs = db.execute(
                "SELECT pr.quantity_required, i.name as item_name, i.unit "
                "FROM protocol_requirements pr JOIN inventory i ON pr.item_id = i.id "
                "WHERE pr.protocol_id = ?", (t_dict['protocol_id'],)
            ).fetchall()
            t_dict['protocol_requirements'] = [dict(req) for req in reqs]
            
            result.append(t_dict)
            
        return jsonify(result)
        
    # POST (Supervisor or Coordinator only)
    if current_user['role'] not in ['coordinator', 'supervisor']:
        return jsonify({'message': 'Forbidden'}), 403
        
    data = request.json or {}
    protocol_id = data.get('protocol_id')
    worker_id = data.get('worker_id')
    node_id = data.get('node_id')
    photo_location_b64 = data.get('photo_location')
    
    if not protocol_id or not worker_id or not node_id:
        return jsonify({'message': 'Missing protocol, worker, or facility node'}), 400
        
    fn_location = save_base64_image(photo_location_b64, 'location') if photo_location_b64 else None
    
    db.execute(
        "INSERT INTO tasks (protocol_id, worker_id, supervisor_id, node_id, status, photo_location) VALUES (?, ?, ?, ?, 'assigned', ?)",
        (protocol_id, worker_id, current_user['id'], node_id, fn_location)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'CREATE_TASK', f"Assigned protocol {protocol_id} on node {node_id} to worker {worker_id}")
    db.commit()
    return jsonify({'message': 'Task created and assigned successfully'})


@app.route('/api/tasks/<int:task_id>', methods=['PUT', 'DELETE'])
@login_required(['coordinator', 'supervisor'])
def task_edit_delete(task_id):
    db = get_db()
    current_user = get_current_user()
    
    # Check if task exists
    task = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return jsonify({'message': 'Task not found'}), 404
        
    if request.method == 'DELETE':
        db.execute("DELETE FROM tasks WHERE id = ?", (task_id,))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'DELETE_TASK', f"Deleted task {task_id}")
        db.commit()
        return jsonify({'message': 'Task deleted successfully'})
        
    # PUT
    data = request.json or {}
    protocol_id = data.get('protocol_id')
    worker_id = data.get('worker_id')
    node_id = data.get('node_id')
    status = data.get('status')
    notes = data.get('notes')
    photo_location = data.get('photo_location')
    
    if not protocol_id or not worker_id or not node_id:
        return jsonify({'message': 'Missing protocol, worker, or facility node'}), 400
        
    final_photo_location = task['photo_location']
    if photo_location:
        # Check if it is a new base64 string or an existing image path
        if photo_location.startswith('data:image') or ',' in photo_location or len(photo_location) > 200:
            fn = save_base64_image(photo_location, 'location')
            if fn:
                final_photo_location = fn
        else:
            final_photo_location = photo_location
    elif 'photo_location' in data and data['photo_location'] is None:
        final_photo_location = None
        
    db.execute(
        "UPDATE tasks SET protocol_id = ?, worker_id = ?, node_id = ?, status = ?, notes = ?, photo_location = ? WHERE id = ?",
        (protocol_id, worker_id, node_id, status or task['status'], notes or task['notes'], final_photo_location, task_id)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'EDIT_TASK', f"Edited task {task_id}")
    db.commit()
    return jsonify({'message': 'Task updated successfully'})

@app.route('/api/tasks/<int:task_id>/status', methods=['POST'])
@login_required(['worker'])
def task_status_update(task_id):
    db = get_db()
    current_user = get_current_user()
    
    data = request.json or {}
    status = data.get('status')
    
    task = db.execute("SELECT * FROM tasks WHERE id = ? AND worker_id = ?", (task_id, current_user['id'])).fetchone()
    if not task:
        return jsonify({'message': 'Task not found or unauthorized'}), 404
        
    now = datetime.datetime.now().isoformat()
    
    if status == 'accepted' and task['status'] in ['assigned', 'rejected']:
        db.execute("UPDATE tasks SET status = 'accepted' WHERE id = ?", (task_id,))
    elif status == 'in_progress' and task['status'] == 'accepted':
        photo_before_b64 = data.get('photo_before')
        fn_before = save_base64_image(photo_before_b64, 'before') if photo_before_b64 else None
        final_before = fn_before if fn_before else task['photo_before']
        db.execute("UPDATE tasks SET status = 'in_progress', start_time = ?, photo_before = ? WHERE id = ?", (now, final_before, task_id))
    elif status == 'accepted' and task['status'] == 'in_progress':
        db.execute("UPDATE tasks SET status = 'accepted' WHERE id = ?", (task_id,))
    else:
        return jsonify({'message': 'Invalid status transition'}), 400
        
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'UPDATE_TASK_STATUS', f"Task #{task_id} status updated to {status}")
    db.commit()
    return jsonify({'message': f'Task status updated to {status}'})

# save_base64_image is now imported from utils.image_utils (with security validation)

@app.route('/api/tasks/<int:task_id>/submit', methods=['POST'])
@login_required(['worker'])
def task_submit(task_id):
    db = get_db()
    current_user = get_current_user()
    
    task = db.execute("SELECT * FROM tasks WHERE id = ? AND worker_id = ?", (task_id, current_user['id'])).fetchone()
    if not task:
        return jsonify({'message': 'Task not found'}), 404
        
    data = request.json or {}
    notes = data.get('notes')
    consumptions = data.get('consumptions', [])
    photo_before_b64 = data.get('photo_before')
    photo_after_b64 = data.get('photo_after')
    
    fn_before = save_base64_image(photo_before_b64, 'before')
    fn_after = save_base64_image(photo_after_b64, 'after')
    
    final_before = fn_before if fn_before else task['photo_before']
    final_after = fn_after if fn_after else task['photo_after']
    
    now = datetime.datetime.now().isoformat()
    
    db.execute("DELETE FROM task_consumptions WHERE task_id = ?", (task_id,))
    
    for c in consumptions:
        item_id = c['item_id']
        qty_used = float(c['quantity'])
        
        db.execute("UPDATE inventory SET stock = stock - ? WHERE id = ?", (qty_used, item_id))
        
        db.execute(
            "INSERT INTO task_consumptions (task_id, item_id, quantity_used) VALUES (?, ?, ?)",
            (task_id, item_id, qty_used)
        )
        
        db.execute(
            "INSERT INTO inventory_logs (item_id, task_id, quantity, user_id, notes) VALUES (?, ?, ?, ?, ?)",
            (item_id, task_id, -qty_used, current_user['id'], f"Consumed in task #{task_id}")
        )
        
    db.execute(
        "UPDATE tasks SET status = 'pending_validation', notes = ?, end_time = ?, photo_before = ?, photo_after = ? WHERE id = ?",
        (notes, now, final_before, final_after, task_id)
    )
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'SUBMIT_TASK', f"Task #{task_id} completed and submitted for validation")
    db.commit()
    return jsonify({'message': 'Task submitted for validation successfully'})

@app.route('/api/tasks/<int:task_id>/validate', methods=['POST'])
@login_required(['coordinator', 'supervisor', 'validator'])
def task_validate(task_id):
    db = get_db()
    current_user = get_current_user()
    
    task = db.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
    if not task:
        return jsonify({'message': 'Task not found'}), 404
        
    data = request.json or {}
    approved = data.get('approved')
    rejection_reason = data.get('rejection_reason', '')
    
    if approved:
        db.execute("UPDATE tasks SET status = 'completed', rejection_reason = NULL WHERE id = ?", (task_id,))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'VALIDATE_TASK_APPROVE', f"Task #{task_id} approved by validation")
        db.commit()
        return jsonify({'message': 'Task approved successfully'})
    else:
        if not rejection_reason:
            return jsonify({'message': 'Rejection reason is required'}), 400
            
        db.execute("UPDATE tasks SET status = 'rejected', rejection_reason = ? WHERE id = ?", (rejection_reason, task_id))
        log_audit(current_user['id'], current_user['username'], current_user['role'], 'VALIDATE_TASK_REJECT', f"Task #{task_id} rejected. Reason: {rejection_reason}")
        db.commit()
        return jsonify({'message': 'Task rejected and returned to worker'})

# --- Reports & Analytics API ---
@app.route('/api/reports', methods=['GET'])
@login_required(['coordinator', 'supervisor'])
def reports_data():
    db = get_db()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    query = (
        "SELECT t.*, p.name as protocol_name, fn.name as node_name, w.name as worker_name, s.name as supervisor_name "
        "FROM tasks t "
        "JOIN protocols p ON t.protocol_id = p.id "
        "JOIN facility_nodes fn ON t.node_id = fn.id "
        "JOIN users w ON t.worker_id = w.id "
        "JOIN users s ON t.supervisor_id = s.id "
        "WHERE t.status = 'completed'"
    )
    params = []
    if start_date:
        query += " AND date(t.end_time) >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date(t.end_time) <= ?"
        params.append(end_date)
        
    completed_tasks = db.execute(query, params).fetchall()
    
    task_durations = []
    worker_times = {}
    machine_times = {}
    supervisor_counts = {}
    
    for t in completed_tasks:
        supervisor_counts[t['supervisor_name']] = supervisor_counts.get(t['supervisor_name'], 0) + 1
        
        if t['start_time'] and t['end_time']:
            try:
                start = datetime.datetime.fromisoformat(t['start_time'])
                end = datetime.datetime.fromisoformat(t['end_time'])
                duration_min = (end - start).total_seconds() / 60.0
                task_durations.append(duration_min)
                
                worker_times.setdefault(t['worker_name'], []).append(duration_min)
                machine_times.setdefault(t['node_name'], []).append(duration_min)
            except Exception:
                pass
                
    avg_task_duration = sum(task_durations) / len(task_durations) if task_durations else 0.0
    
    worker_perf = []
    for w_name, times in worker_times.items():
        worker_perf.append({
            'name': w_name,
            'avg_time': round(sum(times) / len(times), 1),
            'tasks_completed': len(times)
        })
        
    machine_perf = []
    for m_name, times in machine_times.items():
        machine_perf.append({
            'name': m_name,
            'avg_time': round(sum(times) / len(times), 1),
            'tasks_completed': len(times)
        })
        
    supervisor_perf = []
    for s_name, count in supervisor_counts.items():
        supervisor_perf.append({
            'name': s_name,
            'tasks_assigned': count
        })
        
    inv_rows = db.execute("SELECT * FROM inventory WHERE is_active = 1").fetchall()
    inventory_data = []
    low_stock_alerts = 0
    for i in inv_rows:
        is_low = i['stock'] < i['min_stock']
        if is_low:
             low_stock_alerts += 1
        inventory_data.append({
             'id': i['id'],
             'name': i['name'],
             'category': i['category'],
             'stock': i['stock'],
             'min_stock': i['min_stock'],
             'unit': i['unit'],
             'is_low': is_low
        })
        
    counts_query = "SELECT status, COUNT(*) as count FROM tasks WHERE 1=1"
    counts_params = []
    if start_date:
        counts_query += " AND date(created_at) >= ?"
        counts_params.append(start_date)
    if end_date:
        counts_query += " AND date(created_at) <= ?"
        counts_params.append(end_date)
    counts_query += " GROUP BY status"
    counts = db.execute(counts_query, counts_params).fetchall()
    
    status_counts = {c['status']: c['count'] for c in counts}
    for s in ['assigned', 'accepted', 'in_progress', 'pending_validation', 'completed', 'rejected']:
         status_counts.setdefault(s, 0)
         
    app_query = "SELECT COUNT(*) as count FROM audit_logs WHERE action = 'VALIDATE_TASK_APPROVE'"
    rej_query = "SELECT COUNT(*) as count FROM audit_logs WHERE action = 'VALIDATE_TASK_REJECT'"
    val_params = []
    if start_date:
        app_query += " AND date(created_at) >= ?"
        rej_query += " AND date(created_at) >= ?"
        val_params.append(start_date)
    if end_date:
        app_query += " AND date(created_at) <= ?"
        rej_query += " AND date(created_at) <= ?"
        val_params.append(end_date)
        
    approvals = db.execute(app_query, val_params).fetchone()['count']
    rejections = db.execute(rej_query, val_params).fetchone()['count']
    total_validations = approvals + rejections
    validation_rate = (approvals / total_validations * 100) if total_validations > 0 else 100.0
    
    recent_query = (
         "SELECT il.*, i.name as item_name, i.unit "
         "FROM inventory_logs il JOIN inventory i ON il.item_id = i.id "
         "WHERE il.quantity < 0"
    )
    recent_params = []
    if start_date:
        recent_query += " AND date(il.created_at) >= ?"
        recent_params.append(start_date)
    if end_date:
        recent_query += " AND date(il.created_at) <= ?"
        recent_params.append(end_date)
    recent_query += " ORDER BY il.created_at DESC LIMIT 10"
    
    recent_consumptions = db.execute(recent_query, recent_params).fetchall()
    
    return jsonify({
        'avg_task_duration': round(avg_task_duration, 1),
        'worker_performance': worker_perf,
        'machine_performance': machine_perf,
        'supervisor_performance': supervisor_perf,
        'inventory': inventory_data,
        'low_stock_alerts': low_stock_alerts,
        'status_counts': status_counts,
        'validation_rate': round(validation_rate, 1),
        'recent_consumptions': [dict(rc) for rc in recent_consumptions]
    })

# --- Audit API ---
@app.route('/api/audit', methods=['GET'])
@login_required(['coordinator'])
def audit_logs():
    db = get_db()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    username = request.args.get('username')
    
    query = "SELECT * FROM audit_logs WHERE 1=1"
    params = []
    
    if start_date:
        query += " AND date(created_at) >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date(created_at) <= ?"
        params.append(end_date)
    if username:
        query += " AND username = ?"
        params.append(username)
        
    query += " ORDER BY created_at DESC LIMIT 500"
    
    rows = db.execute(query, params).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/api/audit/export', methods=['GET'])
@login_required(['coordinator'])
def export_audit_logs():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from flask import send_file
    
    db = get_db()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    username = request.args.get('username')
    
    query = "SELECT * FROM audit_logs WHERE 1=1"
    params = []
    
    if start_date:
        query += " AND date(created_at) >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date(created_at) <= ?"
        params.append(end_date)
    if username:
        query += " AND username = ?"
        params.append(username)
        
    query += " ORDER BY created_at DESC"
    
    rows = db.execute(query, params).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    headers = ["Timestamp", "Actor", "Role", "Action Code", "Details"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    for r in rows:
        ws.append([r['created_at'], r['username'], r['role'].capitalize(), r['action'], r['details']])
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c_cell = ws.cell(row=curr_row, column=col_idx)
            c_cell.font = normal_font
            c_cell.border = thin_border
            
    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"Audit_Trail_{timestamp}.xlsx"
    
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/reports/export', methods=['GET'])
@login_required(['coordinator', 'supervisor'])
def export_reports_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from flask import send_file
    
    db = get_db()
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # 1. Gather all metrics (same logic as /api/reports endpoint)
    query = (
        "SELECT t.*, p.name as protocol_name, fn.name as node_name, w.name as worker_name, s.name as supervisor_name "
        "FROM tasks t "
        "JOIN protocols p ON t.protocol_id = p.id "
        "JOIN facility_nodes fn ON t.node_id = fn.id "
        "JOIN users w ON t.worker_id = w.id "
        "JOIN users s ON t.supervisor_id = s.id "
        "WHERE t.status = 'completed'"
    )
    params = []
    if start_date:
        query += " AND date(t.end_time) >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date(t.end_time) <= ?"
        params.append(end_date)
        
    completed_tasks = db.execute(query, params).fetchall()
    
    task_durations = []
    worker_times = {}
    machine_times = {}
    supervisor_counts = {}
    
    for t in completed_tasks:
        supervisor_counts[t['supervisor_name']] = supervisor_counts.get(t['supervisor_name'], 0) + 1
        
        if t['start_time'] and t['end_time']:
            try:
                start = datetime.datetime.fromisoformat(t['start_time'])
                end = datetime.datetime.fromisoformat(t['end_time'])
                duration_min = (end - start).total_seconds() / 60.0
                task_durations.append(duration_min)
                worker_times.setdefault(t['worker_name'], []).append(duration_min)
                machine_times.setdefault(t['node_name'], []).append(duration_min)
            except Exception:
                pass
                
    avg_task_duration = sum(task_durations) / len(task_durations) if task_durations else 0.0
    
    # Worker performance
    worker_perf = []
    for w_name, times in worker_times.items():
        worker_perf.append({
            'name': w_name,
            'avg_time': round(sum(times) / len(times), 1),
            'tasks_completed': len(times)
        })
        
    # Machine performance
    machine_perf = []
    for m_name, times in machine_times.items():
        machine_perf.append({
            'name': m_name,
            'avg_time': round(sum(times) / len(times), 1),
            'tasks_completed': len(times)
        })
        
    # Supervisor performance
    supervisor_perf = []
    for s_name, count in supervisor_counts.items():
        supervisor_perf.append({
            'name': s_name,
            'tasks_assigned': count
        })
        
    # Inventory Stock Counts (needed for Low Stock Alerts metric in Overview)
    inv_rows = db.execute("SELECT * FROM inventory WHERE is_active = 1").fetchall()
    low_stock_alerts = 0
    for i in inv_rows:
        is_low = i['stock'] < i['min_stock']
        if is_low:
             low_stock_alerts += 1
        
    # Tasks status count
    counts_query = "SELECT status, COUNT(*) as count FROM tasks WHERE 1=1"
    counts_params = []
    if start_date:
        counts_query += " AND date(created_at) >= ?"
        counts_params.append(start_date)
    if end_date:
        counts_query += " AND date(created_at) <= ?"
        counts_params.append(end_date)
    counts_query += " GROUP BY status"
    counts = db.execute(counts_query, counts_params).fetchall()
    
    status_counts = {c['status']: c['count'] for c in counts}
    for s in ['assigned', 'accepted', 'in_progress', 'pending_validation', 'completed', 'rejected']:
          status_counts.setdefault(s, 0)
          
    # Validation rates
    app_query = "SELECT COUNT(*) as count FROM audit_logs WHERE action = 'VALIDATE_TASK_APPROVE'"
    rej_query = "SELECT COUNT(*) as count FROM audit_logs WHERE action = 'VALIDATE_TASK_REJECT'"
    val_params = []
    if start_date:
        app_query += " AND date(created_at) >= ?"
        rej_query += " AND date(created_at) >= ?"
        val_params.append(start_date)
    if end_date:
        app_query += " AND date(created_at) <= ?"
        rej_query += " AND date(created_at) <= ?"
        val_params.append(end_date)
        
    approvals = db.execute(app_query, val_params).fetchone()['count']
    rejections = db.execute(rej_query, val_params).fetchone()['count']
    total_validations = approvals + rejections
    validation_rate = (approvals / total_validations * 100) if total_validations > 0 else 100.0
    
    # 2. Build the openpyxl workbook
    wb = openpyxl.Workbook()
    
    # Define styles
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    title_font = Font(name="Calibri", size=16, bold=True, color="1F2937")
    section_font = Font(name="Calibri", size=13, bold=True, color="1F2937")
    bold_font = Font(name="Calibri", size=11, bold=True)
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    # Sheet 1: Overview
    ws_ov = wb.active
    ws_ov.title = "Overview"
    ws_ov.views.sheetView[0].showGridLines = True
    
    ws_ov.append([])
    ws_ov.append(["Sanitation System Analytics Report"])
    ws_ov.cell(row=2, column=1).font = title_font
    ws_ov.append([f"Generated at: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    ws_ov.cell(row=3, column=1).font = Font(italic=True, size=10)
    ws_ov.append([])
    
    ws_ov.append(["Key Metrics"])
    ws_ov.cell(row=5, column=1).font = section_font
    
    metrics = [
        ("Average Task Completion Time", f"{round(avg_task_duration, 1)} mins"),
        ("Validation Quality Approval Rate", f"{round(validation_rate, 1)}%"),
        ("Active Low Stock Alerts Count", f"{low_stock_alerts} products"),
    ]
    for m_label, m_val in metrics:
        ws_ov.append([m_label, m_val])
        curr_row = ws_ov.max_row
        ws_ov.cell(row=curr_row, column=1).font = normal_font
        ws_ov.cell(row=curr_row, column=2).font = bold_font
        ws_ov.cell(row=curr_row, column=1).border = thin_border
        ws_ov.cell(row=curr_row, column=2).border = thin_border
        
    ws_ov.append([])
    ws_ov.append(["Tasks Breakdown by Status"])
    ws_ov.cell(row=10, column=1).font = section_font
    
    ws_ov.append(["Status", "Count"])
    for col_idx in (1, 2):
        cell = ws_ov.cell(row=11, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="left")
        
    for status_name, count_val in status_counts.items():
        ws_ov.append([status_name.replace('_', ' ').capitalize(), count_val])
        curr_row = ws_ov.max_row
        ws_ov.cell(row=curr_row, column=1).font = normal_font
        ws_ov.cell(row=curr_row, column=2).font = normal_font
        ws_ov.cell(row=curr_row, column=1).border = thin_border
        ws_ov.cell(row=curr_row, column=2).border = thin_border
        
    # Helper function to auto-adjust column width and style tables
    def format_sheet_table(ws, headers):
        ws.views.sheetView[0].showGridLines = True
        # Style headers
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        # Borders for data
        for row in range(2, ws.max_row + 1):
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.font = normal_font
                cell.border = thin_border
        # Auto width
        for col in ws.columns:
            max_len = 0
            for cell in col:
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    # Sheet 2: Worker Performance
    ws_wp = wb.create_sheet(title="Worker Performance")
    ws_wp.append(["Worker Name", "Average Time (mins)", "Tasks Completed"])
    for wp in worker_perf:
        ws_wp.append([wp['name'], wp['avg_time'], wp['tasks_completed']])
    format_sheet_table(ws_wp, ["Worker Name", "Average Time (mins)", "Tasks Completed"])
    
    # Sheet 3: Machine Performance
    ws_mp = wb.create_sheet(title="Machine Performance")
    ws_mp.append(["Machine/Area", "Average Time (mins)", "Tasks Completed"])
    for mp in machine_perf:
        ws_mp.append([mp['name'], mp['avg_time'], mp['tasks_completed']])
    format_sheet_table(ws_mp, ["Machine/Area", "Average Time (mins)", "Tasks Completed"])
    
    # Sheet 4: Supervisor Performance
    ws_sp = wb.create_sheet(title="Supervisor Activity")
    ws_sp.append(["Supervisor Name", "Tasks Created/Assigned"])
    for sp in supervisor_perf:
        ws_sp.append([sp['name'], sp['tasks_assigned']])
    format_sheet_table(ws_sp, ["Supervisor Name", "Tasks Created/Assigned"])
    
    # Adjust Overview sheet width
    for col in ws_ov.columns:
        max_len = 0
        for cell in col:
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws_ov.column_dimensions[col_letter].width = max(max_len + 4, 15)
        
    # Save workbook to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"Sanitation_Report_{timestamp}.xlsx"
    
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/inventory/export', methods=['GET'])
@login_required(['coordinator', 'supervisor'])
def export_inventory_catalog():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from flask import send_file
    
    db = get_db()
    inv_rows = db.execute("SELECT * FROM inventory WHERE is_active = 1").fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory Catalog"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    headers = ["Product Name", "Category", "Current Stock", "Alert Threshold", "Unit", "Status"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    for i in inv_rows:
        is_low = i['stock'] < i['min_stock']
        status_label = "LOW STOCK ALERT" if is_low else "Optimal"
        ws.append([i['name'], i['category'].capitalize(), i['stock'], i['min_stock'], i['unit'], status_label])
        curr_row = ws.max_row
        
        # Apply red color for LOW STOCK status
        status_cell = ws.cell(row=curr_row, column=6)
        if is_low:
            status_cell.font = Font(name="Calibri", size=11, bold=True, color="FF0000")
        else:
            status_cell.font = normal_font
            
        for col_idx in range(1, len(headers) + 1):
            c_cell = ws.cell(row=curr_row, column=col_idx)
            c_cell.border = thin_border
            if col_idx != 6 or not is_low:
                c_cell.font = normal_font
                
    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"Inventory_Catalog_{timestamp}.xlsx"
    
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/inventory/logs/export', methods=['GET'])
@login_required(['coordinator', 'supervisor'])
def export_consumption_logs():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from flask import send_file
    
    db = get_db()
    
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    username = request.args.get('username')
    
    query = "SELECT il.*, i.name as item_name, i.unit, u.username as user_name " \
            "FROM inventory_logs il " \
            "JOIN inventory i ON il.item_id = i.id " \
            "LEFT JOIN users u ON il.user_id = u.id " \
            "WHERE il.quantity < 0"
    params = []
    
    if start_date:
        query += " AND date(il.created_at) >= ?"
        params.append(start_date)
    if end_date:
        query += " AND date(il.created_at) <= ?"
        params.append(end_date)
    if username:
        query += " AND u.username = ?"
        params.append(username)
        
    query += " ORDER BY il.created_at DESC"
    
    rows = db.execute(query, params).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Consumption Logs"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    headers = ["Timestamp", "Product Name", "Quantity Consumed", "Unit", "Task ID", "Logged By", "Notes"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    for r in rows:
        ws.append([r['created_at'], r['item_name'], abs(r['quantity']), r['unit'], r['task_id'], r['user_name'] or 'System', r['notes']])
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c_cell = ws.cell(row=curr_row, column=col_idx)
            c_cell.font = normal_font
            c_cell.border = thin_border
            
    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d')
    filename = f"Consumption_Logs_{timestamp}.xlsx"
    
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/tasks/export', methods=['GET'])
@login_required(['coordinator', 'supervisor', 'validator'])
def export_tasks_excel():
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side
    from flask import send_file
    
    db = get_db()
    current_user = get_current_user()
    
    # Query all tasks with detailed joins
    query = """
        SELECT t.*, p.name as protocol_name, w.name as worker_name, s.name as supervisor_name, fn.name as node_name
        FROM tasks t
        JOIN protocols p ON t.protocol_id = p.id
        JOIN users w ON t.worker_id = w.id
        JOIN users s ON t.supervisor_id = s.id
        JOIN facility_nodes fn ON t.node_id = fn.id
        ORDER BY t.id DESC
    """
    rows = db.execute(query).fetchall()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sanitation Tasks"
    ws.views.sheetView[0].showGridLines = True
    
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    normal_font = Font(name="Calibri", size=11)
    
    thin_border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB')
    )
    
    headers = ["Task ID", "Protocol", "Area/Machine", "Worker", "Supervisor", "Status", "Start Time", "End Time", "Notes", "Rejection Reason"]
    ws.append(headers)
    
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        
    for r in rows:
        ws.append([
            r['id'],
            r['protocol_name'],
            r['node_name'],
            r['worker_name'],
            r['supervisor_name'],
            r['status'],
            r['start_time'] or '',
            r['end_time'] or '',
            r['notes'] or '',
            r['rejection_reason'] or ''
        ])
        curr_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            c_cell = ws.cell(row=curr_row, column=col_idx)
            c_cell.font = normal_font
            c_cell.border = thin_border
            
    # Auto-adjust column width
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)
    
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"Sanitation_Tasks_{timestamp}.xlsx"
    
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'EXPORT_TASKS', 'Exported sanitation tasks to Excel')
    db.commit()
    
    return send_file(
        file_stream,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


@app.route('/api/backup', methods=['POST'])
@login_required(['coordinator'])
def trigger_manual_backup():
    db = get_db()
    current_user = get_current_user()
    backup_db()
    log_audit(current_user['id'], current_user['username'], current_user['role'], 'MANUAL_BACKUP', 'Coordinator triggered database backup')
    db.commit()
    return jsonify({'message': 'Backup created successfully'})

# --- Application Startup ---
if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=5000, debug=True)
